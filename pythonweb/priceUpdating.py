from bs4 import BeautifulSoup
import requests
import urllib.request
import pandas as pd
import time
import xlsxwriter
import openpyxl
from selenium.common.exceptions import *


df = pd.read_excel('project_mercedes.xlsx', usecols="N")
links = df['URUNLINK'].tolist()
rows = [i for i in range(1, len(links))]

row = 2
temp = 0
tic = 0
i = 0

headers = {"user-agent" : 'Mozilla/5.0 (Macintosh; Intel Mac OS X 10_15_7) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/90.0.4430.212 Safari/537.36'}

productPrices = []
productLinks = []

site = "https://www.onlineyedekparca.com/kategori/mercedes-yedek-parca"  ### CHANGE
page = requests.get(site, headers=headers)
soup = BeautifulSoup(page.content, 'html.parser')

try:
    pages = soup.find(class_ = 'paginate-content')
    pageLinks = pages.find_all('a', href=True)
    pageNumbers = pages.get_text().strip().split('\n')
except :
    pageLinks = []
    pageLinks.append("NoPage")
print( pageNumbers )
for pg in range( 1, int(pageNumbers[5])+1 ):
    if pg != 1:
        linkToGet = "https://www.onlineyedekparca.com/kategori/mercedes-yedek-parca?tp={}".format(pg)  ##* CHANGE
        page = requests.get(linkToGet, headers=headers)
        soup = BeautifulSoup(page.content, 'html.parser')
    
    products = soup.find_all(class_ = 'showcase-content')
    for product in products:
        productLink = product.find(class_ = 'showcase-title')
        productLink = productLink.find('a', href=True)['href']
        price_two = product.find(class_ = 'showcase-price-new').get_text().strip() ## current price
        try:
            price_one = product.find(class_ = 'showcase-price-old').get_text().strip() ## if discounted this is old price
            dummy = price_two
            price_two = price_one
            price_one = dummy
        except :
            price_one = price_two
        price_two = price_two.split(" ")[0].replace(".", "") ## old price
        price_one = price_one.split(" ")[0].replace(".", "") ## new price
        productLinks.append(productLink)
        productPrices.append([float(price_two.replace(',','.')), float(price_one.replace(',','.'))])

    if i == temp*10:
        if i != 0:
            print("] took: {} sec.   {}/{}\n".format((time.time() - tic), pg, int(pageNumbers[5])+1)) ##* Change
        print("{} -> {} [".format(pg, pg+10), end="")
        tic = time.time()
        temp += 1

    print(">", end="")
    i += 1

exceptions = []
writings = [] 
writings.append(productPrices)
writings.append(productLinks)

## Openpyxl
book = openpyxl.load_workbook('project_mercedes.xlsx')  ##* CHANGE
sheet = book.active

suma = 1
for item in links:
    item_solo = item.replace("https://www.onlineyedekparca.com", "")
    if item_solo in writings[1]:
        suma += 1
        idx = writings[1].index(item_solo)
        sheet.cell(row=row , column=14).value = writings[1][idx]
        sheet.cell(row=row, column=11).value = writings[0][idx][0]
        sheet.cell(row=row, column=12).value = writings[0][idx][1]
    else:
        exceptions.append(item)
        sheet.cell(row=row, column=11).value = 0
        sheet.cell(row=row, column=12).value = 0
    row += 1

textfile = open("kaldırılan_ürünler_mercedes_october.txt", "w") ##* change
for element in exceptions:
    textfile.write(element + "\n")
textfile.close()
    
print(suma)
book.save("mercedes_updated_october.xlsx")  ##* change